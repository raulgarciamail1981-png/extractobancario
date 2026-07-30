"""La carpeta de subidas es una cola compartida, no "lo que acabo de subir".

/unify procesa todos los archivos que haya, vengan de donde vengan, y los que
no se pueden leer se quedan ahí reintentándose. Sin verlo, el resultado parece
sumar movimientos "de la nada".
"""
import json

import pytest
from openpyxl import Workbook
from werkzeug.security import generate_password_hash

import db
import web_app


def _extracto(path, cuenta, filas):
    wb = Workbook()
    ws = wb.active
    ws.append(['Cuenta Nro.', cuenta])
    ws.append(['Fecha', 'Descripcion', 'Monto', 'Saldo'])
    for fecha, desc, monto, saldo in filas:
        ws.append([fecha, desc, monto, saldo])
    wb.save(path)


@pytest.fixture
def app_env(tmp_path, monkeypatch):
    upload_dir = tmp_path / 'uploads'
    upload_dir.mkdir()
    users = [
        # Quitar archivos de la cola es un permiso individual, no del rol.
        {'username': 'admin', 'password': generate_password_hash('clave12345'),
         'roles': ['admin'], 'puede_quitar_de_cola': True},
        {'username': 'sin_permiso', 'password': generate_password_hash('clave12345'),
         'roles': ['admin']},
        {'username': 'finanzas_habilitada', 'password': generate_password_hash('clave12345'),
         'roles': ['finanzas'], 'puede_quitar_de_cola': True},
    ]
    (tmp_path / 'users.json').write_text(json.dumps({'users': users}), encoding='utf-8')

    # Maestros mínimos: sin ellos la cuenta no se resuelve y /unify manda todo
    # a la pantalla de "Resolver faltantes" en vez de unificar.
    wb = Workbook()
    ws = wb.active
    ws.append(['CUIT', 'Empresa'])
    ws.append(['20111111116', 'EMPRESA TEST SA'])
    wb.save(tmp_path / 'Empresas.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.append(['BANCO', 'MONEDA', 'EMPRESA TEST SA CUIT 20-11111111-6'])
    ws.append(['SANTANDER', 'PESOS', '1112223334'])
    wb.save(tmp_path / 'DATOS BANCARIOS TODAS LAS EMPRESAS.xlsx')

    monkeypatch.setattr(web_app, 'BASE_DIR', tmp_path)
    monkeypatch.setattr(web_app, 'DB_PATH', tmp_path / 'test.db')
    monkeypatch.setattr(web_app, 'UPLOAD_DIR', upload_dir)
    monkeypatch.setattr(web_app, 'USERS_FILE', tmp_path / 'users.json')
    monkeypatch.setattr(web_app, 'EMPRESAS_PATH', tmp_path / 'Empresas.xlsx')
    monkeypatch.setattr(web_app, 'users_by_name', {u['username']: u for u in users})
    monkeypatch.setattr(web_app, '_login_failures', {})
    web_app.app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
    return {'db_path': tmp_path / 'test.db', 'uploads': upload_dir}


@pytest.fixture
def client(app_env):
    with web_app.app.test_client() as c:
        c.post('/login', data={'username': 'admin', 'password': 'clave12345'})
        yield c


def test_la_pantalla_de_carga_lista_los_archivos_en_cola(client, app_env):
    # Sin esto no hay forma de saber que quedó algo de una subida anterior.
    _extracto(app_env['uploads'] / 'de_ayer.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento viejo', '100,00', '900,00')])

    body = client.get('/upload').get_data(as_text=True)

    assert 'de_ayer.xlsx' in body
    assert '1 archivo(s) en cola' in body


def test_sin_archivos_lo_dice_explicitamente(client, app_env):
    body = client.get('/upload').get_data(as_text=True)

    assert 'No hay archivos en cola' in body


def test_cada_archivo_en_cola_trae_su_boton_de_quitar(client, app_env):
    _extracto(app_env['uploads'] / 'extracto.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento', '100,00', '900,00')])

    body = client.get('/upload').get_data(as_text=True)

    assert '/upload/quitar' in body
    assert 'name="archivo" value="extracto.xlsx"' in body


def test_quitar_saca_el_archivo_sin_unificar_nada(client, app_env):
    # El caso que lo motivó: alguien sube el maestro Empresas.xlsx por error.
    # No aporta movimientos, así que /unify nunca lo borra y queda para siempre.
    _extracto(app_env['uploads'] / 'Empresas.xlsx', '1112223334', [])
    _extracto(app_env['uploads'] / 'extracto.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento', '100,00', '900,00')])

    body = client.post('/upload/quitar', data={'archivo': 'Empresas.xlsx'}).get_data(as_text=True)

    assert 'Se quitó' in body
    assert not (app_env['uploads'] / 'Empresas.xlsx').exists()
    assert (app_env['uploads'] / 'extracto.xlsx').exists(), 'no debe tocar los demás'
    assert len(db.load_movements(db_path=app_env['db_path'])) == 0, 'no debe unificar nada'


def test_quitar_queda_registrado_en_la_auditoria(client, app_env):
    _extracto(app_env['uploads'] / 'extracto.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento', '100,00', '900,00')])

    client.post('/upload/quitar', data={'archivo': 'extracto.xlsx'})

    entradas = db.load_audit_log(db_path=app_env['db_path'])
    assert entradas[0]['action'] == 'upload_delete'
    assert entradas[0]['detail']['archivo'] == 'extracto.xlsx'


@pytest.mark.parametrize('intento', [
    '../users.json', '..\\users.json', '/etc/passwd', 'no_existe.xlsx', '',
])
def test_quitar_no_puede_apuntar_fuera_de_la_cola(client, app_env, intento):
    # El nombre llega de un formulario: no puede servir para borrar cualquier
    # archivo de la máquina.
    (app_env['uploads'].parent / 'users.json').write_text('{"users": []}', encoding='utf-8')
    _extracto(app_env['uploads'] / 'extracto.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento', '100,00', '900,00')])

    body = client.post('/upload/quitar', data={'archivo': intento}).get_data(as_text=True)

    assert 'No se encontró' in body
    assert (app_env['uploads'].parent / 'users.json').exists()
    assert (app_env['uploads'] / 'extracto.xlsx').exists()


def test_un_rol_sin_permiso_de_carga_no_puede_quitar(app_env, tmp_path, monkeypatch):
    usuarios = [{'username': 'viewer', 'password': generate_password_hash('clave12345'),
                 'roles': ['viewer'], 'puede_quitar_de_cola': True}]
    monkeypatch.setattr(web_app, 'users_by_name', {u['username']: u for u in usuarios})
    _extracto(app_env['uploads'] / 'extracto.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento', '100,00', '900,00')])

    with web_app.app.test_client() as c:
        c.post('/login', data={'username': 'viewer', 'password': 'clave12345'})
        resp = c.post('/upload/quitar', data={'archivo': 'extracto.xlsx'})

    # Ni con el permiso individual: un viewer no entra a la pantalla de Carga.
    assert resp.status_code == 403
    assert (app_env['uploads'] / 'extracto.xlsx').exists()


def test_sin_el_permiso_individual_no_se_puede_quitar(app_env):
    # Mismo rol admin, pero sin "puede_quitar_de_cola" en users.json.
    _extracto(app_env['uploads'] / 'extracto.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento', '100,00', '900,00')])

    with web_app.app.test_client() as c:
        c.post('/login', data={'username': 'sin_permiso', 'password': 'clave12345'})
        resp = c.post('/upload/quitar', data={'archivo': 'extracto.xlsx'})

    assert resp.status_code == 403
    assert (app_env['uploads'] / 'extracto.xlsx').exists(), 'el archivo no debe borrarse'


def test_sin_el_permiso_el_boton_ni_se_muestra(app_env):
    _extracto(app_env['uploads'] / 'extracto.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento', '100,00', '900,00')])

    with web_app.app.test_client() as c:
        c.post('/login', data={'username': 'sin_permiso', 'password': 'clave12345'})
        body = c.get('/upload').get_data(as_text=True)

    assert 'extracto.xlsx' in body, 'la cola se sigue viendo'
    assert '/upload/quitar' not in body, 'el botón no debe aparecer'


def test_una_finanzas_habilitada_si_puede_quitar(app_env):
    _extracto(app_env['uploads'] / 'extracto.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento', '100,00', '900,00')])

    with web_app.app.test_client() as c:
        c.post('/login', data={'username': 'finanzas_habilitada', 'password': 'clave12345'})
        body = c.post('/upload/quitar', data={'archivo': 'extracto.xlsx'}).get_data(as_text=True)

    assert 'Se quitó' in body
    assert not (app_env['uploads'] / 'extracto.xlsx').exists()


def test_las_tres_habilitadas_del_padron_real_tienen_el_permiso():
    # Contrato con lo pedido: solo María González, Melani Oviedo y Nicol
    # Feldman pueden quitar archivos de la cola.
    from pathlib import Path
    archivo = Path(web_app.__file__).resolve().parent / 'users.json'
    if not archivo.exists():
        pytest.skip('users.json no está en esta máquina')
    usuarios = json.loads(archivo.read_text(encoding='utf-8'))['users']
    habilitados = {u['username'] for u in usuarios if u.get('puede_quitar_de_cola')}

    assert habilitados == {'magonzalez', 'moviedo', 'nfeldman'}


def test_unificar_procesa_tambien_lo_que_quedo_de_antes(client, app_env):
    # El comportamiento que sorprende: subo un archivo, unifico, y entran
    # también los movimientos de un archivo que había quedado de antes.
    _extracto(app_env['uploads'] / 'de_ayer.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento viejo', '100,00', '900,00')])
    _extracto(app_env['uploads'] / 'de_hoy.xlsx', '1112223334',
              [('16/07/2026', 'Movimiento nuevo', '200,00', '1100,00')])

    body = client.post('/unify').get_data(as_text=True)

    assert 'de_ayer.xlsx' in body and 'de_hoy.xlsx' in body
    assert len(db.load_movements(db_path=app_env['db_path'])) == 2


def test_un_archivo_ilegible_se_informa_y_no_desaparece(client, app_env):
    _extracto(app_env['uploads'] / 'bueno.xlsx', '1112223334',
              [('15/07/2026', 'Movimiento', '100,00', '900,00')])
    (app_env['uploads'] / 'roto.csv').write_bytes(b'esto no es un extracto\n')

    body = client.post('/unify').get_data(as_text=True)

    assert 'no aportaron movimientos' in body
    assert 'roto.csv' in body
    # Sigue en la carpeta: el usuario tiene que poder sacarlo o corregirlo.
    assert (app_env['uploads'] / 'roto.csv').exists()
    assert not (app_env['uploads'] / 'bueno.xlsx').exists()


def test_reunificar_dos_veces_el_mismo_extracto_no_duplica(client, app_env):
    filas = [('15/07/2026', 'Movimiento', '100,00', '900,00'),
             ('16/07/2026', 'Otro', '200,00', '1100,00')]
    _extracto(app_env['uploads'] / 'extracto.xlsx', '1112223334', filas)
    client.post('/unify')
    _extracto(app_env['uploads'] / 'extracto.xlsx', '1112223334', filas)

    client.post('/unify')

    assert len(db.load_movements(db_path=app_env['db_path'])) == 2


def test_dos_extractos_de_la_misma_cuenta_con_rangos_solapados(client, app_env):
    # El caso real: un export del 17 al 22 y otro del 20 al 28. Lo que queda
    # es la unión, no la suma ni el más grande de los dos.
    _extracto(app_env['uploads'] / 'del_17_al_22.xlsx', '1112223334', [
        ('17/07/2026', 'Mov A', '100,00', '1100,00'),
        ('20/07/2026', 'Mov B', '200,00', '1300,00'),
        ('22/07/2026', 'Mov C', '300,00', '1600,00'),
    ])
    _extracto(app_env['uploads'] / 'del_20_al_28.xlsx', '1112223334', [
        ('20/07/2026', 'Mov B', '200,00', '1300,00'),
        ('22/07/2026', 'Mov C', '300,00', '1600,00'),
        ('28/07/2026', 'Mov D', '400,00', '2000,00'),
    ])

    client.post('/unify')

    guardados = db.load_movements(db_path=app_env['db_path'])
    assert len(guardados) == 4, 'se esperaba la unión: A, B, C, D'
    # La descripción se arma con varias columnas segun el banco; alcanza con
    # que estén los cuatro movimientos y ninguno repetido.
    for esperado in ('Mov A', 'Mov B', 'Mov C', 'Mov D'):
        assert sum(esperado in d for d in guardados['Descripcion']) == 1, esperado
