import io
import json
import re
import time
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader
from werkzeug.security import check_password_hash, generate_password_hash

import db
import web_app
from conciliador import hash_record

# Carpeta real del proyecto: web_app.BASE_DIR queda apuntando al tmp_path del
# fixture, así que no sirve para leer templates ni assets versionados.
PROJECT_DIR = Path(web_app.__file__).resolve().parent


@pytest.fixture
def app_env(tmp_path, monkeypatch):
    db_path = tmp_path / 'test.db'
    upload_dir = tmp_path / 'uploads'
    upload_dir.mkdir()
    users_file = tmp_path / 'users.json'
    users = [
        {'username': 'admin', 'password': generate_password_hash('adminpass'), 'role': 'admin'},
        {'username': 'cajera', 'password': generate_password_hash('cajerapass'), 'role': 'cajera'},
    ]
    users_file.write_text(json.dumps({'users': users}), encoding='utf-8')

    # Por defecto no existe Empresas.xlsx en el entorno de test: el resumen
    # de CI cae al listado de empresas derivado de los datos cargados (mismo
    # comportamiento que antes de tener un listado fijo).
    empresas_path = tmp_path / 'Empresas.xlsx'

    # BASE_DIR también se aísla: /unify busca ahí "DATOS BANCARIOS TODAS LAS
    # EMPRESAS.xlsx". Sin esto, los tests leen la carpeta real del proyecto y
    # el resultado depende de si esa máquina tiene el archivo de la empresa.
    monkeypatch.setattr(web_app, 'BASE_DIR', tmp_path)
    monkeypatch.setattr(web_app, 'DB_PATH', db_path)
    monkeypatch.setattr(web_app, 'UPLOAD_DIR', upload_dir)
    monkeypatch.setattr(web_app, 'USERS_FILE', users_file)
    monkeypatch.setattr(web_app, 'EMPRESAS_PATH', empresas_path)
    monkeypatch.setattr(web_app, 'users_by_name', {u['username']: u for u in users})

    # El throttle de login vive en memoria del módulo: sin limpiarlo, los
    # fallos de un test bloquean el login del siguiente.
    monkeypatch.setattr(web_app, '_login_failures', {})

    # CSRF apagado por defecto para no tener que pedir un token en cada POST de
    # los tests. Que la protección esté realmente activa se verifica aparte, en
    # los tests que la encienden a propósito.
    web_app.app.config.update(TESTING=True, WTF_CSRF_ENABLED=False)
    return {
        'db_path': db_path, 'upload_dir': upload_dir, 'users_file': users_file,
        'empresas_path': empresas_path,
        'bank_accounts_path': tmp_path / 'DATOS BANCARIOS TODAS LAS EMPRESAS.xlsx',
    }


@pytest.fixture
def client(app_env):
    with web_app.app.test_client() as test_client:
        yield test_client


def _seed_movement(db_path, record_hash='hash1', ci=''):
    record = {
        'RecordHash': record_hash, 'Fecha': '15/07/2026', 'Empresa': 'Empresa Test',
        'CUIT': '20111111116', 'Cuenta': '1112223334', 'Moneda': 'ARS', 'Banco': 'Banco Test',
        'Descripcion': 'Movimiento', 'Debito': None, 'Credito': 100.0, 'Monto': 100.0,
        'Saldo': 1000.0, 'CI': ci, 'SourceFile': 'archivo1.xlsx', 'SourceRow': 1,
    }
    db.upsert_movements(pd.DataFrame([record]), db_path=db_path)


def login(client, username, password):
    return client.post('/login', data={'username': username, 'password': password}, follow_redirects=True)


def test_records_requires_login(client):
    response = client.get('/records')
    assert response.status_code == 302
    assert response.headers['Location'].rstrip('/') in ('', 'http://localhost')


def test_invalid_login_logs_failure(client, app_env):
    response = client.post('/login', data={'username': 'admin', 'password': 'wrong'})
    assert 'incorrectos'.encode('utf-8') in response.data
    entries = db.load_audit_log(db_path=app_env['db_path'])
    assert entries[0]['action'] == 'login_failed'


def test_valid_login_redirects_and_logs_success(client, app_env):
    response = login(client, 'admin', 'adminpass')
    assert response.status_code == 200
    entries = db.load_audit_log(db_path=app_env['db_path'])
    assert any(e['action'] == 'login_success' for e in entries)


def test_role_without_permission_shows_no_access_without_logging_out(client, app_env):
    login(client, 'cajera', 'cajerapass')

    response = client.get('/upload')

    assert response.status_code == 403
    assert 'No tenés permiso' in response.get_data(as_text=True)
    # La sesión sigue activa: una ruta permitida para su rol debe seguir andando.
    still_logged_in = client.get('/records')
    assert still_logged_in.status_code == 200


def test_records_hides_carga_link_for_roles_without_upload_access(client, app_env):
    _seed_movement(app_env['db_path'])
    login(client, 'cajera', 'cajerapass')

    body = client.get('/records').get_data(as_text=True)

    assert 'href="/upload"' not in body


def test_records_shows_carga_link_for_roles_with_upload_access(client, app_env):
    _seed_movement(app_env['db_path'])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    assert 'href="/upload"' in body


def test_ci_update_preserves_and_audits_change(client, app_env):
    _seed_movement(app_env['db_path'], record_hash='hash1', ci='')
    login(client, 'cajera', 'cajerapass')

    response = client.post('/records', data={'record_hash': 'hash1', 'ci_hash1': 'CI-500'})

    assert 'CI-500' in response.get_data(as_text=True)
    loaded = db.load_movements(db_path=app_env['db_path'])
    assert loaded.iloc[0]['CI'] == 'CI-500'
    entries = db.load_audit_log(db_path=app_env['db_path'])
    assert entries[0]['action'] == 'update_ci'
    assert entries[0]['detail']['despues'] == 'CI-500'


def _seed_movements(db_path, records):
    rows = []
    for i, overrides in enumerate(records, start=1):
        record = {
            'RecordHash': f'hash{i}', 'Fecha': '15/07/2026', 'Empresa': 'Empresa Test',
            'CUIT': '20111111116', 'Cuenta': '1112223334', 'Moneda': 'ARS', 'Banco': 'Banco Test',
            'Descripcion': 'Movimiento', 'Debito': None, 'Credito': 100.0, 'Monto': 100.0,
            'Saldo': 1000.0, 'CI': '', 'SourceFile': 'archivo1.xlsx', 'SourceRow': i,
        }
        record.update(overrides)
        rows.append(record)
    db.upsert_movements(pd.DataFrame(rows), db_path=db_path)


def test_records_filter_dropdowns_list_known_companies_and_banks(client, app_env):
    _seed_movements(app_env['db_path'], [
        {'Empresa': 'Hikari SA', 'Banco': 'Galicia'},
        {'Empresa': 'Alco Rosario SA', 'Banco': 'Santander'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    assert '<option value="Hikari SA"' in body
    assert '<option value="Alco Rosario SA"' in body
    assert '<option value="Galicia"' in body
    assert '<option value="Santander"' in body


def test_records_empresa_and_banco_dropdowns_cascade_with_other_filters(client, app_env):
    # Bug real reportado: al filtrar por Moneda=USD, los desplegables de
    # Empresa y Banco seguían mostrando TODAS las empresas/bancos de la base,
    # no solo los que realmente tienen movimientos en esa moneda.
    _seed_movements(app_env['db_path'], [
        {'Empresa': 'Hikari SA', 'Banco': 'Santa Fe', 'Moneda': 'USD'},
        {'Empresa': 'Alco Rosario SA', 'Banco': 'Santander', 'Moneda': 'ARS'},
        {'Empresa': 'Daseos SA', 'Banco': 'Macro', 'Moneda': 'ARS'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records?moneda=USD').get_data(as_text=True)

    assert '<option value="Hikari SA"' in body
    assert '<option value="Alco Rosario SA"' not in body
    assert '<option value="Daseos SA"' not in body
    assert '<option value="Santa Fe"' in body
    assert '<option value="Santander"' not in body
    assert '<option value="Macro"' not in body


def test_records_empresa_dropdown_excludes_only_its_own_filter(client, app_env):
    # El desplegable de Empresa debe reflejar el filtro de Banco activo, pero
    # no auto-restringirse por el propio filtro de Empresa ya elegido (si no,
    # nunca se podría cambiar de una empresa a otra dentro del mismo banco).
    _seed_movements(app_env['db_path'], [
        {'Empresa': 'Hikari SA', 'Banco': 'Santa Fe'},
        {'Empresa': 'Alco Rosario SA', 'Banco': 'Santa Fe'},
        {'Empresa': 'Daseos SA', 'Banco': 'Macro'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records?banco=Santa+Fe&empresa=Hikari+SA').get_data(as_text=True)

    assert '<option value="Hikari SA"' in body
    assert '<option value="Alco Rosario SA"' in body
    assert '<option value="Daseos SA"' not in body


def test_records_ci_filter_shows_only_empty_or_only_filled(client, app_env):
    _seed_movements(app_env['db_path'], [
        {'CI': 'CI-1'},
        {'CI': ''},
    ])
    login(client, 'admin', 'adminpass')

    only_empty = client.get('/records?ci_filter=vacios').get_data(as_text=True)
    only_filled = client.get('/records?ci_filter=con_ci').get_data(as_text=True)

    assert 'CI-1' not in only_empty
    assert 'CI-1' in only_filled


def test_records_general_search_matches_amount_regardless_of_sign(client, app_env):
    _seed_movements(app_env['db_path'], [
        {'Descripcion': 'Pago proveedor', 'Monto': -1500.50},
        {'Descripcion': 'Cobro cliente', 'Monto': 999.00},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records?search=1500,50').get_data(as_text=True)

    assert 'Pago proveedor' in body
    assert 'Cobro cliente' not in body


def test_records_general_search_matches_description_text(client, app_env):
    _seed_movements(app_env['db_path'], [
        {'Descripcion': 'Transferencia recibida de Juan'},
        {'Descripcion': 'Comision bancaria'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records?search=transferencia').get_data(as_text=True)

    assert 'Transferencia recibida de Juan' in body
    assert 'Comision bancaria' not in body


def test_records_general_search_handles_special_characters_literally(client, app_env):
    _seed_movements(app_env['db_path'], [
        {'Descripcion': 'Pago a proveed acred santander rio - &&(0001933)'},
        {'Descripcion': 'Cobro cliente'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records?search=%26%26(0001933)').get_data(as_text=True)

    assert 'santander rio' in body
    assert 'Cobro cliente' not in body


def test_records_general_search_numeric_term_also_matches_description_code(client, app_env):
    # Con los nuevos formatos de descripción (Macro, Santander, etc.) las
    # descripciones incluyen códigos numéricos (sucursal, referencia). Un
    # término numérico debe poder encontrar esos códigos, no solo montos.
    _seed_movements(app_env['db_path'], [
        {'Descripcion': '0718 | Rosario Centro | 4633 | 000105255 | Impuesto ley', 'Monto': -508594.75},
        {'Descripcion': 'Comision bancaria', 'Monto': 12.00},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records?search=000105255').get_data(as_text=True)

    assert 'Rosario Centro' in body
    assert 'Comision bancaria' not in body


def test_records_moneda_filter_shows_only_selected_currency(client, app_env):
    _seed_movements(app_env['db_path'], [
        {'Descripcion': 'Movimiento en pesos', 'Moneda': ''},
        {'Descripcion': 'Movimiento en dolares', 'Moneda': 'USD'},
    ])
    login(client, 'admin', 'adminpass')

    only_usd = client.get('/records?moneda=USD').get_data(as_text=True)
    only_pesos = client.get('/records?moneda=%24').get_data(as_text=True)

    assert 'Movimiento en dolares' in only_usd
    assert 'Movimiento en pesos' not in only_usd
    assert 'Movimiento en pesos' in only_pesos
    assert 'Movimiento en dolares' not in only_pesos


def test_records_date_range_filter_correct_for_ambiguous_day_and_month(client, app_env):
    # "2026-07-08" (día 8, mes 7) es el caso exacto que dayfirst=True
    # interpretaba mal como 7 de agosto. Filtrando desde/hasta esa misma
    # fecha, el movimiento tiene que aparecer.
    _seed_movements(app_env['db_path'], [
        {'Fecha': '08/07/2026', 'Descripcion': 'Movimiento del 8 de julio'},
        {'Fecha': '20/08/2026', 'Descripcion': 'Movimiento de agosto'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records?fecha_desde=2026-07-08&fecha_hasta=2026-07-08').get_data(as_text=True)

    assert 'Movimiento del 8 de julio' in body
    assert 'Movimiento de agosto' not in body


def test_api_latest_unify_reflects_last_unify_action(client, app_env):
    login(client, 'admin', 'adminpass')
    db.log_action('admin', 'unify', detail={'filas_procesadas': 5, 'nuevas': 3}, db_path=app_env['db_path'])

    response = client.get('/api/latest-unify')
    data = response.get_json()

    assert data['username'] == 'admin'
    assert data['detail']['nuevas'] == 3


def test_unify_removes_processed_files_from_uploads(client, app_env):
    # Si el archivo ya unificado se queda en /uploads, la próxima
    # unificación lo vuelve a procesar junto con lo nuevo que se suba.
    wb = Workbook()
    ws = wb.active
    ws.append(['Fecha', 'Descripcion', 'Monto', 'Saldo', 'Empresa', 'CUIT'])
    ws.append(['15/07/2026', 'Movimiento de prueba', '100,00', '900,00', 'Empresa Test', '20111111116'])
    target = app_env['upload_dir'] / 'archivo_test_unify.xlsx'
    wb.save(target)

    login(client, 'admin', 'adminpass')
    response = client.post('/unify', data={}, follow_redirects=True)

    assert response.status_code == 200
    assert not target.exists()
    body = response.get_data(as_text=True)
    assert 'se quitaron de la carpeta de subidas' in body
    assert 'Archivos procesados' in body
    assert 'archivo_test_unify.xlsx: 1 filas leídas, 1 nuevas.' in body
    assert 'Total de filas leídas: 1.' in body
    assert 'Nuevas agregadas: 1.' in body
    assert len(db.load_movements(db_path=app_env['db_path'])) == 1


def test_unify_reports_per_file_new_vs_repeated_rows(client, app_env):
    # Dos archivos con fechas superpuestas (como dos exports de ICBC en
    # distintos momentos): el segundo debe mostrar cuántas de sus filas ya
    # habían entrado con el primero.
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.append(['Fecha', 'Descripcion', 'Monto', 'Saldo', 'Empresa', 'CUIT'])
    ws1.append(['01/07/2026', 'Movimiento A', '100,00', '900,00', 'Empresa Test', '20111111116'])
    ws1.append(['02/07/2026', 'Movimiento B', '200,00', '700,00', 'Empresa Test', '20111111116'])
    wb1.save(app_env['upload_dir'] / 'archivo_1_primero.xlsx')

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(['Fecha', 'Descripcion', 'Monto', 'Saldo', 'Empresa', 'CUIT'])
    ws2.append(['02/07/2026', 'Movimiento B', '200,00', '700,00', 'Empresa Test', '20111111116'])
    ws2.append(['03/07/2026', 'Movimiento C', '300,00', '400,00', 'Empresa Test', '20111111116'])
    wb2.save(app_env['upload_dir'] / 'archivo_2_segundo.xlsx')

    login(client, 'admin', 'adminpass')
    response = client.post('/unify', data={}, follow_redirects=True)

    body = response.get_data(as_text=True)
    assert 'archivo_1_primero.xlsx: 2 filas leídas, 2 nuevas.' in body
    assert 'archivo_2_segundo.xlsx: 2 filas leídas, 1 nuevas (1 ya existían).' in body
    assert 'Total de filas leídas: 4.' in body
    assert 'Nuevas agregadas: 3 (1 ya existían y no se duplicaron).' in body
    assert len(db.load_movements(db_path=app_env['db_path'])) == 3


def _seed_macro_btob(db_path, record_hash, nro_referencia, monto, empresa='Alco Rosario SA',
                      cuenta='376109405439550', fecha='08/05/2026'):
    record = {
        'RecordHash': record_hash, 'Fecha': fecha, 'Empresa': empresa,
        'CUIT': '20111111116', 'Cuenta': cuenta, 'Moneda': 'ARS', 'Banco': 'Macro',
        'Descripcion': f'TEF DATANET BTOB | {nro_referencia} | 2026', 'Debito': abs(monto),
        'Credito': None, 'Monto': monto, 'Saldo': 1000000.0, 'CI': '', 'SourceFile': 'archivo1.xls',
        'SourceRow': 1, 'RAW_nro de referencia': nro_referencia,
    }
    db.upsert_movements(pd.DataFrame([record]), db_path=db_path)


def test_result_shows_interbanking_prompt_only_when_pending(client, app_env):
    login(client, 'admin', 'adminpass')

    # Sin nada pendiente todavía: unificar un extracto común no debe preguntar.
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.append(['Fecha', 'Descripcion', 'Monto', 'Saldo', 'Empresa', 'CUIT'])
    ws1.append(['01/07/2026', 'Movimiento A', '100,00', '900,00', 'Empresa Test', '20111111116'])
    wb1.save(app_env['upload_dir'] / 'archivo1.xlsx')
    response = client.post('/unify', data={}, follow_redirects=True)
    assert 'Unificar registros Interbanking' not in response.get_data(as_text=True)

    # Con un movimiento Macro "TEF DATANET BTOB" pendiente en la base, el
    # próximo /unify que unifique algo nuevo debe preguntar.
    _seed_macro_btob(app_env['db_path'], 'btob1', '120378926', -2163345.13)
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(['Fecha', 'Descripcion', 'Monto', 'Saldo', 'Empresa', 'CUIT'])
    ws2.append(['05/07/2026', 'Movimiento B', '200,00', '700,00', 'Empresa Test', '20111111116'])
    wb2.save(app_env['upload_dir'] / 'archivo2.xlsx')
    response2 = client.post('/unify', data={}, follow_redirects=True)
    assert 'Unificar registros Interbanking' in response2.get_data(as_text=True)


def test_interbanking_requires_upload_role(client, app_env):
    login(client, 'cajera', 'cajerapass')

    response = client.get('/interbanking')

    assert response.status_code == 403


def test_interbanking_lists_candidates_for_default_pending_account(client, app_env):
    _seed_macro_btob(app_env['db_path'], 'btob1', '120378926', -2163345.13)
    _seed_macro_btob(app_env['db_path'], 'btob2', '120378928', -375933.48)
    login(client, 'admin', 'adminpass')

    body = client.get('/interbanking').get_data(as_text=True)

    assert 'TEF DATANET BTOB | 120378926 | 2026' in body
    assert 'TEF DATANET BTOB | 120378928 | 2026' in body
    assert 'Alco Rosario SA' in body


def test_interbanking_pre_checks_exact_amount_suggestion(client, app_env):
    _seed_macro_btob(app_env['db_path'], 'btob1', '111', -55705270.66)
    _seed_macro_btob(app_env['db_path'], 'btob2', '222', -55705270.66, fecha='09/05/2026')
    login(client, 'admin', 'adminpass')

    body = client.get('/interbanking').get_data(as_text=True)

    # El de fecha más reciente (btob2) queda como "A Compensar"; el otro pre-tildado.
    assert 'value="btob2" class="a-compensar-radio" checked' in body
    assert 'value="btob1" class="select-checkbox" checked' in body


def test_interbanking_apply_with_exact_sum_tags_all_records(client, app_env):
    _seed_macro_btob(app_env['db_path'], 'a_compensar', '999999', -1520000.0)
    _seed_macro_btob(app_env['db_path'], 'detalle1', '111', -120000.0)
    _seed_macro_btob(app_env['db_path'], 'detalle2', '222', -1000000.0)
    _seed_macro_btob(app_env['db_path'], 'detalle3', '333', -400000.0)
    login(client, 'admin', 'adminpass')

    response = client.post('/interbanking/apply', data={
        'empresa': 'Alco Rosario SA', 'cuenta': '376109405439550',
        'a_compensar_hash': 'a_compensar',
        'selected_hashes': ['detalle1', 'detalle2', 'detalle3'],
    })

    body = response.get_data(as_text=True)
    assert 'Se unificaron 4 registros' in body
    loaded = db.load_movements(db_path=app_env['db_path']).set_index('RecordHash')
    for h in ['a_compensar', 'detalle1', 'detalle2', 'detalle3']:
        assert loaded.loc[h, 'RegistroUnificado'] == '999999'
    entries = db.load_audit_log(db_path=app_env['db_path'])
    assert any(e['action'] == 'interbanking_unify' for e in entries)


def test_interbanking_apply_with_mismatched_sum_does_not_modify_anything(client, app_env):
    _seed_macro_btob(app_env['db_path'], 'a_compensar', '999999', -1520000.0)
    _seed_macro_btob(app_env['db_path'], 'detalle1', '111', -120000.0)
    login(client, 'admin', 'adminpass')

    response = client.post('/interbanking/apply', data={
        'empresa': 'Alco Rosario SA', 'cuenta': '376109405439550',
        'a_compensar_hash': 'a_compensar',
        'selected_hashes': ['detalle1'],
    })

    body = response.get_data(as_text=True)
    assert 'no coincide' in body
    loaded = db.load_movements(db_path=app_env['db_path']).set_index('RecordHash')
    assert loaded.loc['a_compensar', 'RegistroUnificado'] == ''
    assert loaded.loc['detalle1', 'RegistroUnificado'] == ''


def test_records_shows_registro_unificado_column_and_is_searchable(client, app_env):
    _seed_macro_btob(app_env['db_path'], 'btob1', '999999', -1520000.0)
    db.apply_interbanking_group(['btob1'], '999999', db_path=app_env['db_path'])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)
    assert 'Registro Unificado' in body
    assert '<td data-label="Registro Unificado">999999</td>' in body

    filtered = client.get('/records?search=999999').get_data(as_text=True)
    assert 'TEF DATANET BTOB' in filtered


def test_admin_create_user_stores_nombre(client, app_env):
    login(client, 'admin', 'adminpass')

    response = client.post('/admin', data={
        'action': 'create', 'username': 'nueva', 'password': 'pass12345',
        'role': 'viewer', 'nombre': 'Juana Perez',
    })

    assert response.status_code == 200
    users_after = json.loads(app_env['users_file'].read_text(encoding='utf-8'))
    nueva = next(u for u in users_after['users'] if u['username'] == 'nueva')
    assert nueva['nombre'] == 'Juana Perez'


def test_admin_update_user_preserves_nombre_when_left_blank(client, app_env):
    login(client, 'admin', 'adminpass')
    client.post('/admin', data={
        'action': 'create', 'username': 'nueva', 'password': 'pass12345',
        'role': 'viewer', 'nombre': 'Juana Perez',
    })

    client.post('/admin', data={'action': 'update', 'username': 'nueva', 'role': 'cajera'})

    users_after = json.loads(app_env['users_file'].read_text(encoding='utf-8'))
    nueva = next(u for u in users_after['users'] if u['username'] == 'nueva')
    assert nueva['nombre'] == 'Juana Perez'
    # Los roles se guardan como lista: hay usuarias con más de uno.
    assert nueva['roles'] == ['cajera']


def test_admin_update_does_not_drop_the_second_role(client, app_env):
    # Editar desde el admin a una usuaria finanzas+cajera no puede dejarla con
    # uno solo por el simple hecho de haber tocado otro campo.
    login(client, 'admin', 'adminpass')
    client.post('/admin', data={
        'action': 'create', 'username': 'doble', 'password': 'pass12345',
        'roles': ['finanzas', 'cajera'], 'nombre': 'Finanzas Y Cajera',
    })

    client.post('/admin', data={
        'action': 'update', 'username': 'doble', 'roles': ['finanzas', 'cajera'],
        'nombre': 'Finanzas Y Cajera',
    })

    users_after = json.loads(app_env['users_file'].read_text(encoding='utf-8'))
    doble = next(u for u in users_after['users'] if u['username'] == 'doble')
    assert sorted(doble['roles']) == ['cajera', 'finanzas']


def test_admin_update_preserves_company_assignments(client, app_env):
    # Las empresas asignadas no se editan desde este formulario: tocar el rol
    # de una cajera no puede borrárselas.
    users = json.loads(app_env['users_file'].read_text(encoding='utf-8'))
    users['users'].append({
        'username': 'cajera1', 'nombre': 'Cajera Uno', 'password': 'x', 'roles': ['cajera'],
        'empresas_primarias': ['ALCO'], 'empresas_secundarias': ['HIKARI'],
    })
    app_env['users_file'].write_text(json.dumps(users), encoding='utf-8')
    login(client, 'admin', 'adminpass')

    client.post('/admin', data={'action': 'update', 'username': 'cajera1', 'roles': ['cajera']})

    after = json.loads(app_env['users_file'].read_text(encoding='utf-8'))
    cajera1 = next(u for u in after['users'] if u['username'] == 'cajera1')
    assert cajera1['empresas_primarias'] == ['ALCO']
    assert cajera1['empresas_secundarias'] == ['HIKARI']


def test_is_valid_password_requires_length_and_alphanumeric():
    assert web_app.is_valid_password('abcd1234') is True
    assert web_app.is_valid_password('short1') is False
    assert web_app.is_valid_password('abcdefgh') is True
    assert web_app.is_valid_password('abcd123!') is False
    assert web_app.is_valid_password('abcd 1234') is False


def test_admin_create_user_rejects_weak_password(client, app_env):
    login(client, 'admin', 'adminpass')

    response = client.post('/admin', data={
        'action': 'create', 'username': 'nueva', 'password': 'short1',
        'role': 'viewer', 'nombre': 'Juana Perez',
    })

    assert 'alfanumérica' in response.get_data(as_text=True)
    users_after = json.loads(app_env['users_file'].read_text(encoding='utf-8'))
    assert all(u['username'] != 'nueva' for u in users_after['users'])


def test_admin_create_user_rejects_non_alphanumeric_password(client, app_env):
    login(client, 'admin', 'adminpass')

    response = client.post('/admin', data={
        'action': 'create', 'username': 'nueva', 'password': 'pass1234!',
        'role': 'viewer',
    })

    assert 'alfanumérica' in response.get_data(as_text=True)
    users_after = json.loads(app_env['users_file'].read_text(encoding='utf-8'))
    assert all(u['username'] != 'nueva' for u in users_after['users'])


def test_admin_update_user_rejects_weak_password(client, app_env):
    login(client, 'admin', 'adminpass')
    client.post('/admin', data={
        'action': 'create', 'username': 'nueva', 'password': 'pass12345', 'role': 'viewer',
    })

    response = client.post('/admin', data={'action': 'update', 'username': 'nueva', 'role': 'viewer', 'password': 'short1'})

    assert 'alfanumérica' in response.get_data(as_text=True)


def test_admin_page_shows_form_before_users_table(client, app_env):
    login(client, 'admin', 'adminpass')

    body = client.get('/admin').get_data(as_text=True)

    assert body.index('id="userForm"') < body.index('id="userTable"')


def test_admin_page_has_search_input_and_edit_buttons(client, app_env):
    login(client, 'admin', 'adminpass')

    body = client.get('/admin').get_data(as_text=True)

    assert 'id="userSearch"' in body
    assert 'edit-user-button' in body
    assert 'data-username="admin"' in body


def test_change_password_requires_login(client):
    response = client.get('/change-password')
    assert response.status_code == 302


def test_change_password_success(client, app_env):
    login(client, 'cajera', 'cajerapass')

    response = client.post('/change-password', data={
        'current_password': 'cajerapass', 'new_password': 'nuevapass123', 'confirm_password': 'nuevapass123',
    })

    assert 'actualizada correctamente' in response.get_data(as_text=True)
    # La nueva contraseña ya sirve para loguearse.
    client.get('/logout')
    login_response = login(client, 'cajera', 'nuevapass123')
    assert login_response.status_code == 200


def test_change_password_rejects_wrong_current_password(client, app_env):
    login(client, 'cajera', 'cajerapass')

    response = client.post('/change-password', data={
        'current_password': 'incorrecta', 'new_password': 'nuevapass123', 'confirm_password': 'nuevapass123',
    })

    assert 'no es correcta' in response.get_data(as_text=True)


def test_change_password_rejects_mismatched_confirmation(client, app_env):
    login(client, 'cajera', 'cajerapass')

    response = client.post('/change-password', data={
        'current_password': 'cajerapass', 'new_password': 'nuevapass123', 'confirm_password': 'otradistinta',
    })

    assert 'no coinciden' in response.get_data(as_text=True)


def test_change_password_rejects_weak_new_password(client, app_env):
    login(client, 'cajera', 'cajerapass')

    response = client.post('/change-password', data={
        'current_password': 'cajerapass', 'new_password': 'short1', 'confirm_password': 'short1',
    })

    assert 'alfanumérica' in response.get_data(as_text=True)


def test_records_shows_date_as_dd_mm_yyyy(client, app_env):
    _seed_movement(app_env['db_path'])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    assert '15/07/2026' in body
    assert '2026-07-15' not in body


def test_records_save_ci_button_is_floating_for_cajera(client, app_env):
    _seed_movement(app_env['db_path'])
    login(client, 'cajera', 'cajerapass')

    body = client.get('/records').get_data(as_text=True)

    assert 'save-ci-bar' in body
    assert 'has-save-ci-bar' in body


def test_login_shows_display_name_instead_of_just_role(client, app_env):
    users_file_content = json.loads(app_env['users_file'].read_text(encoding='utf-8'))
    users_file_content['users'][0]['nombre'] = 'Ana Admin'
    app_env['users_file'].write_text(json.dumps(users_file_content), encoding='utf-8')
    import web_app
    web_app.users_by_name['admin']['nombre'] = 'Ana Admin'

    login(client, 'admin', 'adminpass')
    body = client.get('/records').get_data(as_text=True)

    assert 'Ana Admin' in body


def test_admin_clear_wipes_only_movements(client, app_env):
    _seed_movement(app_env['db_path'], record_hash='hash1')
    dummy_upload = app_env['upload_dir'] / 'dummy.xlsx'
    dummy_upload.write_bytes(b'not a real xlsx, just checking it survives')
    login(client, 'admin', 'adminpass')

    response = client.post('/admin/clear', follow_redirects=True)

    assert response.status_code == 200
    assert db.load_movements(db_path=app_env['db_path']).empty
    assert dummy_upload.exists()
    users_after = json.loads(app_env['users_file'].read_text(encoding='utf-8'))
    assert len(users_after['users']) == 2


def test_records_uses_movimiento_label_and_has_no_apply_button(client, app_env):
    _seed_movement(app_env['db_path'])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    assert 'name="monto_filter"' in body
    assert 'Tipo de monto' not in body
    assert 'Aplicar filtros' not in body
    assert 'last30Button' in body
    assert "filtersForm.addEventListener('change'" in body


def test_ci_alert_flag_only_for_overdue_credit_without_ci(client, app_env):
    overdue_date = (datetime.now() - timedelta(days=15)).strftime('%d/%m/%Y')
    recent_date = (datetime.now() - timedelta(days=2)).strftime('%d/%m/%Y')
    _seed_movements(app_env['db_path'], [
        {'Fecha': overdue_date, 'Monto': 500.0, 'Credito': 500.0, 'CI': '', 'Descripcion': 'Acreditacion vieja sin CI'},
        {'Fecha': recent_date, 'Monto': 500.0, 'Credito': 500.0, 'CI': '', 'Descripcion': 'Acreditacion reciente sin CI'},
        {'Fecha': overdue_date, 'Monto': -500.0, 'Debito': 500.0, 'CI': '', 'Descripcion': 'Debito viejo sin CI'},
        {'Fecha': overdue_date, 'Monto': 500.0, 'Credito': 500.0, 'CI': 'CI-1', 'Descripcion': 'Acreditacion vieja con CI'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    # "ci-alert" también aparece una vez en el <style> (la regla CSS); lo que
    # importa es que solo una fila tenga la clase realmente aplicada al input.
    assert body.count('ci-cell-field ci-alert') == 1


def test_records_ci_summary_shows_total_and_per_company_boxes(client, app_env):
    overdue_date = (datetime.now() - timedelta(days=15)).strftime('%d/%m/%Y')
    recent_date = (datetime.now() - timedelta(days=2)).strftime('%d/%m/%Y')
    _seed_movements(app_env['db_path'], [
        {'Empresa': 'Empresa A', 'Fecha': overdue_date, 'Monto': 500.0, 'Credito': 500.0, 'CI': ''},
        {'Empresa': 'Empresa B', 'Fecha': overdue_date, 'Monto': 500.0, 'Credito': 500.0, 'CI': ''},
        {'Empresa': 'Empresa B', 'Fecha': recent_date, 'Monto': 500.0, 'Credito': 500.0, 'CI': ''},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    # Recuadro TOTAL: 3 acreditaciones sin CI, 2 con más de 10 días.
    assert 'ci-summary-total' in body
    assert '>TOTAL<' in body
    assert '<strong>3</strong> acreditaciones sin CI' in body
    assert '<strong>2</strong> de esas con más de 10 días sin CI' in body

    # Recuadro Empresa A: 1 sin CI, 1 vencida.
    assert '>Empresa A<' in body
    assert '<strong>1</strong> acreditaciones sin CI' in body

    # Recuadro Empresa B: 2 sin CI, 1 vencida.
    assert '>Empresa B<' in body
    assert '<strong>2</strong> acreditaciones sin CI' in body
    assert '<strong>1</strong> de esas con más de 10 días sin CI' in body

    # Los links de cada recuadro llevan los filtros correctos para esa empresa.
    assert 'empresa=Empresa+A' in body
    assert 'ci_filter=vacios' in body
    assert 'monto_filter=positivos' in body
    assert 'vencido=1' in body


def test_records_ci_summary_box_click_filters_to_that_company(client, app_env):
    overdue_date = (datetime.now() - timedelta(days=15)).strftime('%d/%m/%Y')
    recent_date = (datetime.now() - timedelta(days=2)).strftime('%d/%m/%Y')
    _seed_movements(app_env['db_path'], [
        {'Empresa': 'Empresa A', 'Fecha': overdue_date, 'Monto': 500.0, 'Credito': 500.0, 'CI': '', 'Descripcion': 'Mov A vencido'},
        {'Empresa': 'Empresa B', 'Fecha': overdue_date, 'Monto': 500.0, 'Credito': 500.0, 'CI': '', 'Descripcion': 'Mov B vencido'},
        {'Empresa': 'Empresa B', 'Fecha': recent_date, 'Monto': 500.0, 'Credito': 500.0, 'CI': '', 'Descripcion': 'Mov B reciente'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get(
        '/records?empresa=Empresa+B&ci_filter=vacios&monto_filter=positivos&vencido=1'
    ).get_data(as_text=True)

    assert 'Mov B vencido' in body
    assert 'Mov A vencido' not in body
    assert 'Mov B reciente' not in body


def test_empresa_dropdown_lists_master_companies_without_movements(client, app_env):
    # Una empresa recién dada de alta en Empresas.xlsx tiene que poder
    # elegirse aunque todavía no haya cargado extractos: si no figura, no hay
    # forma de distinguir "no cargó nada" de "está mal dada de alta".
    _write_empresas_file(app_env['empresas_path'], [
        ('20111111116', 'Hikari SA'),
        ('20222222223', 'Neostar SA'),
    ])
    _seed_movements(app_env['db_path'], [{'Empresa': 'Hikari SA', 'Banco': 'Galicia'}])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    assert '<option value="Neostar SA"' in body
    assert '<option value="Hikari SA"' in body


def test_selecting_a_company_without_movements_shows_an_empty_grid(client, app_env):
    _write_empresas_file(app_env['empresas_path'], [
        ('20111111116', 'Hikari SA'),
        ('20222222223', 'Neostar SA'),
    ])
    _seed_movements(app_env['db_path'], [
        {'Empresa': 'Hikari SA', 'Descripcion': 'Movimiento de Hikari'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records?empresa=Neostar+SA').get_data(as_text=True)

    assert 'No hay movimientos para mostrar' in body
    assert 'Movimiento de Hikari' not in body


def test_records_table_rows_have_zebra_striping(client, app_env):
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    # El CSS vive en una hoja compartida, no embebido en cada template: se
    # verifica que la página la enlace y que la hoja tenga la regla.
    assert 'css/app.css' in body
    stylesheet = (PROJECT_DIR / 'static' / 'css' / 'app.css').read_text(encoding='utf-8')
    assert 'nth-child(even)' in stylesheet


def _write_empresas_file(path, companies):
    wb = Workbook()
    ws = wb.active
    ws.append(['CUIT', 'Empresa'])
    for cuit, empresa in companies:
        ws.append([cuit, empresa])
    wb.save(path)


def test_records_ci_summary_boxes_are_fixed_even_without_pending_ci(client, app_env):
    # El listado de empresas es el fijo de Empresas.xlsx: debe aparecer un
    # recuadro por cada una aunque no tenga movimientos ni pendientes de CI.
    _write_empresas_file(app_env['empresas_path'], [
        ('20111111116', 'Empresa A'),
        ('20222222223', 'Empresa B'),
        ('20333333330', 'Empresa Sin Movimientos'),
    ])
    _seed_movements(app_env['db_path'], [
        {'Empresa': 'Empresa A', 'CI': 'CI-1'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    names = set()
    for line in body.splitlines():
        if 'ci-summary-name' in line:
            names.add(line.strip())
    assert any('Empresa A' in n for n in names)
    assert any('Empresa B' in n for n in names)
    assert any('Empresa Sin Movimientos' in n for n in names)
    # Empresa A no tiene pendientes de CI (el único movimiento ya tiene CI
    # cargado): el recuadro debe seguir mostrándose, en 0.
    assert '>Empresa A<' in body


def test_download_excel_hides_saldo_for_cajera_and_applies_filters(client, app_env):
    _seed_movements(app_env['db_path'], [
        {'Empresa': 'Empresa A', 'Monto': 100.0, 'Credito': 100.0, 'Saldo': 900.0},
        {'Empresa': 'Empresa B', 'Monto': 200.0, 'Credito': 200.0, 'Saldo': 800.0},
    ])
    login(client, 'cajera', 'cajerapass')

    resp = client.get('/download/excel?empresa=Empresa+A')

    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    all_values = [cell for row in ws.iter_rows(values_only=True) for cell in row if cell is not None]
    assert 'Saldo' not in all_values
    assert any('Empresa A' == v for v in all_values)
    assert not any('Empresa B' == v for v in all_values)
    assert any(isinstance(v, str) and 'Filtros aplicados' in v for v in all_values)


def test_download_excel_shows_saldo_for_admin(client, app_env):
    _seed_movement(app_env['db_path'])
    login(client, 'admin', 'adminpass')

    resp = client.get('/download/excel')

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    all_values = [cell for row in ws.iter_rows(values_only=True) for cell in row if cell is not None]
    assert 'Saldo' in all_values


def test_download_pdf_is_real_pdf_and_respects_role_and_filters(client, app_env):
    _seed_movements(app_env['db_path'], [
        {'Empresa': 'Empresa A', 'Descripcion': 'Movimiento de Empresa A', 'Monto': 100.0, 'Credito': 100.0},
        {'Empresa': 'Empresa B', 'Descripcion': 'Movimiento de Empresa B', 'Monto': 200.0, 'Credito': 200.0},
    ])
    login(client, 'cajera', 'cajerapass')

    resp = client.get('/download/pdf?empresa=Empresa+A')

    assert resp.status_code == 200
    assert resp.mimetype == 'application/pdf'
    assert resp.data[:4] == b'%PDF'
    reader = PdfReader(io.BytesIO(resp.data))
    text = '\n'.join(page.extract_text() or '' for page in reader.pages)
    assert 'Filtros aplicados' in text
    assert 'Empresa A' in text or 'Movimiento de Empresa A' in text
    assert 'Empresa B' not in text
    assert 'Saldo' not in text


# --- users.json ausente -------------------------------------------------

def test_load_users_returns_empty_when_file_is_missing(app_env, monkeypatch):
    # Importar la app no puede depender de un archivo que no está en el repo.
    monkeypatch.setattr(web_app, 'USERS_FILE', app_env['users_file'].parent / 'no_existe.json')

    assert web_app.load_users() == []


def test_login_explains_missing_users_file_instead_of_blaming_credentials(client, monkeypatch):
    monkeypatch.setattr(web_app, 'users_by_name', {})

    response = client.post('/login', data={'username': 'admin', 'password': 'adminpass'})

    body = response.get_data(as_text=True)
    assert 'users.json' in body
    assert 'incorrectos' not in body


def test_users_example_file_is_a_valid_template():
    # Es el archivo que alguien va a copiar a users.json en un clone limpio.
    example = json.loads((PROJECT_DIR / 'users.example.json').read_text(encoding='utf-8'))
    users = example['users']

    assert users, 'la plantilla tiene que traer al menos un usuario'
    for user in users:
        assert user['role'] in web_app.USER_ROLES
        # El hash tiene que ser verificable: si no, copiar la plantilla
        # rompe el login con un error de werkzeug en vez de dejar entrar.
        assert check_password_hash(user['password'], 'cambiame123')


# --- resolución de cuentas no registradas -------------------------------

def _write_bank_accounts(path, account='1112223334'):
    # Layout real de "DATOS BANCARIOS TODAS LAS EMPRESAS.xlsx": tabla pivot con
    # banco y moneda en las dos primeras columnas y una columna por empresa.
    wb = Workbook()
    ws = wb.active
    ws.append(['BANCO', 'MONEDA', 'EMPRESA TEST SA CUIT 20-11111111-6'])
    ws.append(['SANTANDER', 'PESOS', account])
    wb.save(path)


def _write_statement_for_unregistered_account(path):
    wb = Workbook()
    ws = wb.active
    ws.append(['Cuenta Nro. 999-888777/6'])
    ws.append(['Fecha', 'Descripcion', 'Monto', 'Saldo'])
    ws.append(['16/07/2026', 'Transferencia recibida', '1500,50', '20000,00'])
    wb.save(path)


def _resolve_form_from(body, cuit='20111111116', banco='SANTANDER'):
    hashes = re.findall(r'name="record_hash_\d+" value="([0-9a-f]+)"', body)
    data = {'resolve_missing': '1', 'missing_count': str(len(hashes))}
    for index, record_hash in enumerate(hashes):
        data[f'record_hash_{index}'] = record_hash
        data[f'mapping_{index}'] = cuit
        data[f'bank_{index}'] = banco
    return hashes, data


def test_unify_completes_after_assigning_an_account_not_in_datos_bancarios(client, app_env):
    # La pantalla de resolución no cambia el número de cuenta, así que el
    # re-chequeo por cuenta volvía a marcar la fila y devolvía al usuario a la
    # misma pantalla para siempre: una cuenta nueva no se podía unificar nunca.
    _write_bank_accounts(app_env['bank_accounts_path'])
    _write_statement_for_unregistered_account(app_env['upload_dir'] / 'extracto_nuevo.xlsx')
    login(client, 'admin', 'adminpass')

    first = client.post('/unify')
    body = first.get_data(as_text=True)
    # Marcador estructural de la pantalla, no su título: el texto visible es
    # cosa del diseño y no debería romper este test si se reescribe.
    assert 'name="resolve_missing"' in body

    hashes, form = _resolve_form_from(body)
    assert hashes

    second = client.post('/unify', data=form)

    assert 'Archivos procesados' in second.get_data(as_text=True)
    stored = db.load_movements(db_path=app_env['db_path'])
    assert len(stored) == 1
    assert stored.iloc[0]['CUIT'] == '20111111116'
    assert stored.iloc[0]['Banco'] == 'SANTANDER'


def test_unify_recomputes_record_hash_after_manual_assignment(client, app_env):
    # El hash describe el contenido de la fila e incluye Banco y CUIT, así que
    # después de asignarlos a mano tiene que cambiar. Si se guardara con el
    # hash viejo, el día que la cuenta se dé de alta en DATOS BANCARIOS y el
    # banco pase a detectarse solo, el mismo movimiento entraría de nuevo.
    _write_bank_accounts(app_env['bank_accounts_path'])
    _write_statement_for_unregistered_account(app_env['upload_dir'] / 'extracto_nuevo.xlsx')
    login(client, 'admin', 'adminpass')

    body = client.post('/unify').get_data(as_text=True)
    original_hashes, form = _resolve_form_from(body)
    client.post('/unify', data=form)

    stored = db.load_movements(db_path=app_env['db_path'])
    assert stored.iloc[0]['RecordHash'] not in original_hashes


def test_reunifying_the_same_file_with_the_same_assignment_does_not_duplicate(client, app_env):
    # El hash recalculado tiene que ser determinístico: mismo archivo + misma
    # asignación manual = mismo movimiento, no uno nuevo.
    _write_bank_accounts(app_env['bank_accounts_path'])
    login(client, 'admin', 'adminpass')

    for _ in range(2):
        _write_statement_for_unregistered_account(app_env['upload_dir'] / 'extracto_nuevo.xlsx')
        body = client.post('/unify').get_data(as_text=True)
        _, form = _resolve_form_from(body)
        assert 'Archivos procesados' in client.post('/unify', data=form).get_data(as_text=True)

    assert len(db.load_movements(db_path=app_env['db_path'])) == 1


def test_recompute_record_hashes_leaves_untouched_rows_alone(app_env):
    # Contracara del test anterior: recalcular una fila que no cambió tiene
    # que dar exactamente el mismo hash, o cada unificación duplicaría todo.
    row = {
        'RecordHash': 'sin_cambios', 'Banco': 'SANTANDER', 'Cuenta': '1112223334',
        'Fecha': pd.Timestamp('2026-07-16'), 'Monto': 1500.5, 'Saldo': 20000.0,
        'CUIT': '20111111116', 'Descripcion': 'Transferencia recibida',
    }
    df = pd.DataFrame([dict(row, RecordHash=hash_record(row))])
    before = df.iloc[0]['RecordHash']

    after = web_app.recompute_record_hashes(df, [before])

    assert after.iloc[0]['RecordHash'] == before


def test_resolve_screen_lists_each_row_only_once(client, app_env):
    # Una fila sin Empresa/CUIT y con cuenta no registrada cae en los dos
    # grupos; se tiene que pedir una sola vez, no dos.
    _write_bank_accounts(app_env['bank_accounts_path'])
    _write_statement_for_unregistered_account(app_env['upload_dir'] / 'extracto_nuevo.xlsx')
    login(client, 'admin', 'adminpass')

    body = client.post('/unify').get_data(as_text=True)

    hashes = re.findall(r'name="record_hash_\d+" value="([0-9a-f]+)"', body)
    assert len(hashes) == 1
    assert len(set(hashes)) == 1


def test_unify_does_not_ask_to_resolve_rows_the_statement_already_identifies(client, app_env):
    # Cuenta no registrada en DATOS BANCARIOS, pero el extracto trae Empresa y
    # CUIT y el nombre del archivo el banco: no hay nada que preguntar.
    _write_bank_accounts(app_env['bank_accounts_path'])
    wb = Workbook()
    ws = wb.active
    ws.append(['Fecha', 'Descripcion', 'Monto', 'Saldo', 'Cuenta', 'Empresa', 'CUIT'])
    ws.append(['16/07/2026', 'Transferencia', '1500,50', '20000,00', '999888777',
               'Empresa Test SA', '20111111116'])
    wb.save(app_env['upload_dir'] / 'extracto_galicia.xlsx')
    login(client, 'admin', 'adminpass')

    response = client.post('/unify')

    assert 'Archivos procesados' in response.get_data(as_text=True)


# --- descarga de Excel sin archivo temporal compartido -------------------

def test_download_excel_does_not_write_a_shared_temp_file(client, app_env):
    # Se generaba en UPLOAD_DIR con un nombre fijo: dos descargas simultáneas
    # se pisaban y una cajera podía terminar con el Excel de un admin (con la
    # columna Saldo que su rol no debe ver).
    _seed_movement(app_env['db_path'])
    login(client, 'admin', 'adminpass')

    resp = client.get('/download/excel')

    assert resp.status_code == 200
    assert list(app_env['upload_dir'].iterdir()) == []


def test_download_excel_keeps_each_role_seeing_its_own_columns(client, app_env):
    # Dos descargas intercaladas, sin recargar nada en el medio: cada una tiene
    # que traer sus propias columnas.
    _seed_movement(app_env['db_path'])

    login(client, 'cajera', 'cajerapass')
    cajera_bytes = client.get('/download/excel').data
    client.get('/logout')
    login(client, 'admin', 'adminpass')
    admin_bytes = client.get('/download/excel').data

    def headers_of(data):
        ws = load_workbook(io.BytesIO(data)).active
        return [c for row in ws.iter_rows(values_only=True) for c in row if c is not None]

    assert 'Saldo' not in headers_of(cajera_bytes)
    assert 'Saldo' in headers_of(admin_bytes)


# --- CSRF ---------------------------------------------------------------

@pytest.fixture
def csrf_client(app_env):
    web_app.app.config.update(WTF_CSRF_ENABLED=True)
    try:
        with web_app.app.test_client() as test_client:
            yield test_client
    finally:
        web_app.app.config.update(WTF_CSRF_ENABLED=False)


def test_post_without_csrf_token_is_rejected(csrf_client):
    response = csrf_client.post('/login', data={'username': 'admin', 'password': 'adminpass'})

    assert response.status_code == 400


def test_post_with_csrf_token_from_the_form_is_accepted(csrf_client, app_env):
    page = csrf_client.get('/').get_data(as_text=True)
    token = re.search(r'name="csrf_token" value="([^"]+)"', page).group(1)

    response = csrf_client.post(
        '/login', data={'username': 'admin', 'password': 'adminpass', 'csrf_token': token},
    )

    assert response.status_code == 302
    assert any(e['action'] == 'login_success' for e in db.load_audit_log(db_path=app_env['db_path']))


@pytest.mark.parametrize('template', [
    'login.html', 'admin.html', 'change_password.html', 'records.html',
    'resolve_missing.html', 'upload.html',
])
def test_every_post_form_carries_a_csrf_token(template):
    # Un formulario sin token no falla al renderizar: falla recién cuando un
    # usuario lo manda. Este test lo detecta antes.
    html = (PROJECT_DIR / 'templates' / template).read_text(encoding='utf-8')

    assert html.count('method="post"') == html.count('csrf_token()')


# --- cookie de sesión ---------------------------------------------------

def test_session_cookie_is_hardened_by_default():
    assert web_app.app.config['SESSION_COOKIE_HTTPONLY'] is True
    assert web_app.app.config['SESSION_COOKIE_SAMESITE'] == 'Lax'
    # El deploy real es HTTPS detrás de Caddy; el opt-out es solo para dev.
    assert web_app.app.config['SESSION_COOKIE_SECURE'] is True


# --- límite de intentos de login ----------------------------------------

def test_login_is_blocked_after_repeated_failures(client, app_env):
    for _ in range(web_app.LOGIN_MAX_ATTEMPTS - 1):
        assert client.post('/login', data={'username': 'admin', 'password': 'mal'}).status_code == 200

    blocked = client.post('/login', data={'username': 'admin', 'password': 'mal'})

    assert blocked.status_code == 429
    assert 'Demasiados intentos' in blocked.get_data(as_text=True)
    actions = [e['action'] for e in db.load_audit_log(db_path=app_env['db_path'])]
    assert 'login_blocked' in actions


def test_throttled_login_rejects_even_the_correct_password(client):
    for _ in range(web_app.LOGIN_MAX_ATTEMPTS):
        client.post('/login', data={'username': 'admin', 'password': 'mal'})

    response = client.post('/login', data={'username': 'admin', 'password': 'adminpass'})

    assert response.status_code == 429
    with client.session_transaction() as sess:
        assert 'username' not in sess


def test_blocked_attempts_do_not_keep_growing_the_audit_log(client, app_env):
    for _ in range(web_app.LOGIN_MAX_ATTEMPTS + 5):
        client.post('/login', data={'username': 'admin', 'password': 'mal'})

    actions = [e['action'] for e in db.load_audit_log(db_path=app_env['db_path'])]
    assert actions.count('login_blocked') == 1


def test_successful_login_clears_previous_failures(client):
    for _ in range(web_app.LOGIN_MAX_ATTEMPTS - 1):
        client.post('/login', data={'username': 'admin', 'password': 'mal'})

    assert client.post('/login', data={'username': 'admin', 'password': 'adminpass'}).status_code == 302

    client.get('/logout')
    for _ in range(web_app.LOGIN_MAX_ATTEMPTS - 1):
        assert client.post('/login', data={'username': 'admin', 'password': 'mal'}).status_code == 200


def test_failures_from_one_user_do_not_block_another(client):
    for _ in range(web_app.LOGIN_MAX_ATTEMPTS):
        client.post('/login', data={'username': 'admin', 'password': 'mal'})

    response = client.post('/login', data={'username': 'cajera', 'password': 'cajerapass'})

    assert response.status_code == 302


def test_old_failures_fall_out_of_the_window(client, monkeypatch):
    for _ in range(web_app.LOGIN_MAX_ATTEMPTS):
        client.post('/login', data={'username': 'admin', 'password': 'mal'})
    assert client.post('/login', data={'username': 'admin', 'password': 'mal'}).status_code == 429

    real_monotonic = time.monotonic
    monkeypatch.setattr(
        web_app.time, 'monotonic', lambda: real_monotonic() + web_app.LOGIN_WINDOW_SECONDS + 1,
    )

    assert client.post('/login', data={'username': 'admin', 'password': 'adminpass'}).status_code == 302



def _client_como(rol, app_env, monkeypatch):
    usuarios = [{'username': 'usuaria', 'password': generate_password_hash('clave12345'),
                 'roles': [rol], 'nombre': 'Usuaria Test'}]
    monkeypatch.setattr(web_app, 'users_by_name', {u['username']: u for u in usuarios})
    c = web_app.app.test_client()
    c.post('/login', data={'username': 'usuaria', 'password': 'clave12345'})
    return c


def test_asistente_comercial_consulta_el_resumen(app_env, monkeypatch):
    # El asistente comercial es de solo consulta: entra al Resumen y a las
    # descargas, igual que un viewer.
    _seed_movement(app_env['db_path'])
    c = _client_como('comercial', app_env, monkeypatch)

    assert c.get('/records').status_code == 200
    assert c.get('/download/excel').status_code == 200


def test_asistente_comercial_no_carga_ni_administra(app_env, monkeypatch):
    c = _client_como('comercial', app_env, monkeypatch)

    assert c.get('/upload').status_code == 403
    assert c.post('/unify', data={}).status_code == 403
    assert c.get('/admin').status_code == 403


def test_el_rol_comercial_se_muestra_como_asistente_comercial(app_env, monkeypatch):
    c = _client_como('comercial', app_env, monkeypatch)

    body = c.get('/records').get_data(as_text=True)

    assert 'asistente comercial' in body


# --------- integracion de Interbanking con esta app (no toca su logica) ---------

def test_el_formulario_de_interbanking_lleva_el_token_csrf(client, app_env):
    # La proteccion CSRF es de esta app, no venia en el repo de origen. Sin el
    # token, "Aplicar unificacion" se rechaza con 400 en produccion. En los
    # tests la proteccion esta apagada, asi que no salta sola: por eso se
    # verifica que el campo este en el formulario.
    _seed_macro_btob(app_env['db_path'], 'btob1', '111', -100.0)
    login(client, 'admin', 'adminpass')

    body = client.get('/interbanking').get_data(as_text=True)

    assert 'name="csrf_token"' in body


@pytest.mark.parametrize('rol, password, deberia_ver', [
    ('admin', 'adminpass', True),
    ('cajera', 'cajerapass', False),
])
def test_el_acceso_a_interbanking_aparece_en_el_menu_segun_el_rol(client, app_env, rol, password, deberia_ver):
    login(client, rol, password)

    body = client.get('/records').get_data(as_text=True)

    assert ('href="/interbanking"' in body) is deberia_ver


@pytest.mark.parametrize('termino', ['751.200,00', '751200,00', '751.200', '751200'])
def test_la_busqueda_encuentra_el_monto_tal_como_se_ve_en_la_grilla(client, app_env, termino):
    # La grilla muestra "751.200,00" y lo natural es copiar de ahi para buscar.
    # El separador de miles quedaba en el termino pero no en el dato, asi que
    # buscar exactamente lo que se ve no encontraba nada.
    _seed_movements(app_env['db_path'], [
        {'Descripcion': 'Transferencia ctas mobile banking', 'Monto': 751200.00},
        {'Descripcion': 'Otro movimiento', 'Monto': 999.00},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get(f'/records?search={termino}').get_data(as_text=True)

    assert 'Transferencia ctas mobile banking' in body
    assert 'Otro movimiento' not in body


def test_la_busqueda_de_un_monto_con_signo_menos_sigue_andando(client, app_env):
    _seed_movements(app_env['db_path'], [
        {'Descripcion': 'Pago proveedor', 'Monto': -1500.50},
        {'Descripcion': 'Cobro cliente', 'Monto': 999.00},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records?search=-1.500,50').get_data(as_text=True)

    assert 'Pago proveedor' in body
    assert 'Cobro cliente' not in body


def test_el_resumen_ordena_las_fechas_de_mayor_a_menor(client, app_env):
    _seed_movements(app_env['db_path'], [
        {'Descripcion': 'La del medio', 'Fecha': '15/07/2026'},
        {'Descripcion': 'La mas vieja', 'Fecha': '01/07/2026'},
        {'Descripcion': 'La mas nueva', 'Fecha': '31/07/2026'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    assert body.index('La mas nueva') < body.index('La del medio') < body.index('La mas vieja')


def test_dentro_de_la_misma_fecha_se_respeta_el_orden_del_extracto(client, app_env):
    # Orden estable: si el banco lista tres movimientos del mismo dia en cierto
    # orden, ordenar por fecha no se lo tiene que dar vuelta.
    _seed_movements(app_env['db_path'], [
        {'Descripcion': 'Primero del dia', 'Fecha': '15/07/2026'},
        {'Descripcion': 'Segundo del dia', 'Fecha': '15/07/2026'},
        {'Descripcion': 'Tercero del dia', 'Fecha': '15/07/2026'},
    ])
    login(client, 'admin', 'adminpass')

    body = client.get('/records').get_data(as_text=True)

    assert body.index('Primero del dia') < body.index('Segundo del dia') < body.index('Tercero del dia')
